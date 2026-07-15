import csv
import ipaddress
import logging
import re
import socket
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urljoin, urlparse
from urllib.robotparser import RobotFileParser

import requests
from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from core import validation as core_validation
from core.logging import configure_logging

from . import schemas

# Emails embedded as plain text (as opposed to mailto: links).
_TEXT_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
# mailto: links — stop at the first character that ends the address / starts
# query params, so "mailto:foo@bar.com?subject=Hi" yields "foo@bar.com".
_MAILTO_RE = re.compile(r"mailto:([^\"'?>\s]+)", re.IGNORECASE)
# Cloudflare's "Email Address Obfuscation" (Scrape Shield) replaces every
# mailto: link and plain-text email with this span instead, so the raw
# address never appears in the HTML at all; see _decode_cf_email for the
# hex+XOR scheme. Always double-quoted — this is server-rendered by
# Cloudflare's edge, not author-written markup with variable quoting.
_CF_EMAIL_RE = re.compile(r'data-cfemail="([0-9a-fA-F]+)"')

# Contact-form detection (see has_contact_form). The opening and closing
# <form> tags are matched separately (rather than one greedy/non-greedy
# "<form>...</form>" pattern) so a form missing its closing tag can still be
# scanned instead of silently skipped — see _extract_form_blocks.
_FORM_OPEN_RE = re.compile(r"<form\b", re.IGNORECASE)
_FORM_CLOSE_RE = re.compile(r"</form\s*>", re.IGNORECASE)

# A message-type field inside a form block: a <textarea> (freeform message
# box) or an <input type="email"> (reply-to address field), in any attribute
# order and with quoted or unquoted attribute values. Either is a strong
# signal the form is for contacting the site rather than e.g. a search box.
_MESSAGE_FIELD_RE = re.compile(
    r"<textarea\b|<input\b[^>]*\btype\s*=\s*['\"]?email\b", re.IGNORECASE
)

# Contact-related keyword text (Polish + English), matched case-insensitively
# via str.lower() against a form block's text — see _form_block_has_signal.
_CONTACT_FORM_KEYWORDS = (
    "kontakt",
    "napisz do nas",
    "wiadomość",
    "formularz",
    "message",
    "contact us",
    "get in touch",
)

# Statuses requests would otherwise auto-follow; handled manually instead so
# each response can be inspected before deciding whether to chase it. The
# robots.txt loader re-validates and follows same-host hops (see
# MAX_REDIRECTS); fetch_page does not follow redirects at all.
_REDIRECT_STATUS_CODES = frozenset({301, 302, 303, 307, 308})

# Explicit allowlist for fetched page content: substring-matched against the
# lower-cased Content-Type header. Narrower than a loose "html" or "text"
# check, which would also accept text/csv, text/plain, etc.
_ALLOWED_CONTENT_TYPES = ("text/html", "application/xhtml+xml")


def _default_resolve_host(host: str) -> list:
    return socket.getaddrinfo(host, None)


def _is_public_host(
    host: str,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> bool:
    """Reject hosts that resolve to private/loopback/link-local/reserved IPs,
    to prevent SSRF via search results or redirects pointing at internal
    infrastructure (cloud metadata endpoints, RFC-1918 ranges, etc.). The
    resolver is injectable (mirrors http_get/sleep_fn) so tests stay
    hermetic instead of doing a real DNS lookup."""
    try:
        infos = resolve_host(host)
    except socket.gaierror:
        return False
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if (
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_reserved
            or ip.is_multicast
        ):
            return False
    return True


def _redirect_target(response: requests.Response, current_url: str) -> Optional[str]:
    """Absolute redirect target for a 3xx response with a Location header, or
    None if this response isn't a redirect we should follow. Requests'
    allow_redirects auto-follow is disabled everywhere we call this, so
    3xx responses reach here instead of being silently chased."""
    if response.status_code not in _REDIRECT_STATUS_CODES:
        return None
    location = response.headers.get("Location")
    if not location:
        return None
    return urljoin(current_url, location)


def validate_scrape_inputs_for_chain(
    contact_email: str, log_file: Optional[Path]
) -> schemas.ScrapeContactsChainConfig:
    """Validate contact-scrape inputs except the results workbook path, which
    isn't known until the chained prospect-search stage produces it."""
    bootstrap_logger = configure_logging(log_file=None)
    bootstrap_logger.info("Validating contact-scrape inputs...")

    resolved_log_file = core_validation.resolve_log_path(
        log_file, schemas.DEFAULT_LOG_DIR
    )

    config = core_validation.validate_config(
        schemas.ScrapeContactsChainConfig,
        {
            "contact_email": contact_email,
            "log_file": resolved_log_file,
        },
        resolved_log_file,
        schemas.DEFAULT_LOG_DIR,
        "contact-scrape inputs",
    )

    logger = configure_logging(log_file=config.log_file)
    logger.info(
        "Validation passed: contact_email=%s log_file=%s about_us_csv_path=%s "
        "email_prefixes_csv_path=%s",
        config.contact_email,
        config.log_file,
        config.about_us_csv_path,
        config.email_prefixes_csv_path,
    )
    return config


def build_user_agent(contact_email: str) -> str:
    return schemas.USER_AGENT_TEMPLATE.format(email=contact_email)


def read_about_us_slugs(csv_path: Path) -> list[str]:
    """Read the url_slug column from about-us-urls.csv, stripping blanks and
    deduplicating while preserving first-seen order (the CSV has repeats such
    as /bok and /faq)."""
    slugs: list[str] = []
    seen: set[str] = set()
    with csv_path.open(newline="", encoding="utf-8") as f:
        for record in csv.DictReader(f):
            slug = (record.get("url_slug") or "").strip()
            if not slug or slug in seen:
                continue
            seen.add(slug)
            slugs.append(slug)
    return slugs


def read_email_prefixes(csv_path: Path) -> frozenset[str]:
    """Read the prefix column from email_prefixes_only.csv into a lower-cased,
    deduplicated allowlist of generic mailbox prefixes (info, kontakt, ...)."""
    prefixes: set[str] = set()
    with csv_path.open(newline="", encoding="utf-8") as f:
        for record in csv.DictReader(f):
            prefix = (record.get("prefix") or "").strip().lower()
            if prefix:
                prefixes.add(prefix)
    return frozenset(prefixes)


def build_candidate_urls(host: str, slugs: list[str]) -> list[str]:
    """Homepage first (footers frequently carry the contact email), then each
    about-us subpage."""
    urls = [f"https://{host}/"]
    urls.extend(f"https://{host}{slug}" for slug in slugs)
    return urls


def load_robots(
    host: str,
    user_agent: str,
    logger: logging.Logger,
    http_get: Callable[..., requests.Response] = requests.get,
    sleep_fn: Callable[[float], None] = time.sleep,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> RobotFileParser:
    """Fetch and parse a host's robots.txt through the shared requests stack
    (so it honours the OS trust store and our User-Agent). Per RFC 9309
    §2.3.1: 401/403 -> disallow all; other 4xx ("Unavailable") -> allow all
    (nothing to respect); 5xx or network errors -> disallow all (the crawler
    MUST assume the site is fully disallowed until robots.txt is reachable
    again)."""
    rp = RobotFileParser()
    if not _is_public_host(host, resolve_host):
        rp.disallow_all = True
        logger.warning(
            "Host %s does not resolve to a public IP; assuming disallow-all "
            "(SSRF guard)",
            host,
        )
        return rp

    current_url = f"https://{host}/robots.txt"
    sleep_fn(schemas.SCRAPE_DELAY_SECONDS)
    hops = 0
    while True:
        try:
            response = http_get(
                current_url,
                headers={"User-Agent": user_agent},
                timeout=schemas.PAGE_FETCH_TIMEOUT,
                allow_redirects=False,
            )
        except requests.RequestException as exc:
            rp.disallow_all = True
            logger.warning(
                "Could not fetch robots.txt for %s (%s); assuming disallow-all",
                host,
                exc,
            )
            return rp

        redirect_url = _redirect_target(response, current_url)
        if redirect_url is None:
            break

        hops += 1
        if hops > schemas.MAX_REDIRECTS:
            rp.disallow_all = True
            logger.warning(
                "Too many redirects fetching robots.txt for %s; assuming "
                "disallow-all",
                host,
            )
            return rp

        redirect_host = urlparse(redirect_url).hostname
        if redirect_host != host:
            rp.disallow_all = True
            logger.warning(
                "robots.txt for %s redirected off-host to %s; not following, "
                "assuming disallow-all",
                host,
                redirect_url,
            )
            return rp
        if not _is_public_host(redirect_host, resolve_host):
            rp.disallow_all = True
            logger.warning(
                "robots.txt redirect for %s to %s does not resolve to a "
                "public IP; assuming disallow-all (SSRF guard)",
                host,
                redirect_url,
            )
            return rp

        logger.info("Following robots.txt redirect for %s -> %s", host, redirect_url)
        current_url = redirect_url

    if response.status_code == 200:
        rp.parse(response.text.splitlines())
        logger.info("Loaded robots.txt for %s", host)
    elif response.status_code in (401, 403):
        rp.disallow_all = True
        logger.info(
            "robots.txt for %s returned %d; treating as disallow-all",
            host,
            response.status_code,
        )
    elif 400 <= response.status_code < 500:
        rp.allow_all = True
        logger.info(
            "No usable robots.txt for %s (status %d); assuming crawling allowed",
            host,
            response.status_code,
        )
    else:
        rp.disallow_all = True
        logger.warning(
            "robots.txt for %s returned %d; assuming disallow-all",
            host,
            response.status_code,
        )
    return rp


def resolve_crawl_delay(rp: RobotFileParser, user_agent: str) -> float:
    """Effective per-request delay: our baseline, raised to the host's
    Crawl-delay directive when that is larger."""
    crawl_delay = rp.crawl_delay(user_agent)
    if crawl_delay is None:
        return float(schemas.SCRAPE_DELAY_SECONDS)
    return max(float(schemas.SCRAPE_DELAY_SECONDS), float(crawl_delay))


def fetch_page(
    url: str,
    user_agent: str,
    logger: logging.Logger,
    http_get: Callable[..., requests.Response] = requests.get,
    timeout: int = schemas.PAGE_FETCH_TIMEOUT,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> Optional[str]:
    """Return HTML text for url, or None (content unavailable — move on) on a
    request error, a redirect, a non-200 status, non-HTML content, an
    oversized body, or a host that isn't public (SSRF guard). Redirects are
    not followed: a candidate contact page that redirects is treated the
    same as one with no content, so the caller moves straight on to the next
    candidate page instead of spending a second request on a destination
    that's often an anti-bot/captcha interstitial anyway (e.g. www.orteo.pl
    redirects every candidate page to captcha.php)."""
    hostname = urlparse(url).hostname
    if hostname is None or not _is_public_host(hostname, resolve_host):
        logger.warning(
            "Host for %s does not resolve to a public IP; skipping (SSRF guard)",
            url,
        )
        return None

    try:
        response = http_get(
            url,
            headers={
                "User-Agent": user_agent,
                "Accept": "text/html,application/xhtml+xml",
            },
            timeout=timeout,
            allow_redirects=False,
        )
    except requests.RequestException as exc:
        logger.info("No content from %s (%s)", url, exc)
        return None

    redirect_url = _redirect_target(response, url)
    if redirect_url is not None:
        logger.info(
            "Redirect %s -> %s; not following, trying next page", url, redirect_url
        )
        return None

    if response.status_code != 200:
        logger.info("No content from %s (status %d)", url, response.status_code)
        return None

    content_type = response.headers.get("Content-Type", "").lower()
    if not any(allowed in content_type for allowed in _ALLOWED_CONTENT_TYPES):
        logger.info("Skipping non-HTML content at %s (%r)", url, content_type)
        return None

    if len(response.text) > schemas.MAX_PAGE_BYTES:
        logger.warning(
            "Response body for %s exceeds %d bytes; skipping",
            url,
            schemas.MAX_PAGE_BYTES,
        )
        return None

    return response.text


def _is_asset_filename(candidate: str) -> bool:
    lowered = candidate.lower()
    return any(lowered.endswith(ext) for ext in schemas.ASSET_EXTENSIONS)


def _prefix_allowed(local_part: str, allowed_prefixes: frozenset[str]) -> bool:
    """Whether an email's local-part matches the allowlist: equal to a prefix, or
    a prefix followed immediately by a separator (info@, info.sales@, info-pl@)."""
    lp = local_part.lower()
    for prefix in allowed_prefixes:
        if lp == prefix or any(
            lp.startswith(prefix + sep) for sep in schemas.EMAIL_PREFIX_SEPARATORS
        ):
            return True
    return False


def _decode_cf_email(hex_str: str) -> Optional[str]:
    """Reverse Cloudflare's obfuscation: the first byte is a single-byte XOR
    key applied to every remaining byte to recover the address. Returns None
    (rather than raising) for anything that doesn't fit the scheme — odd
    length or non-hex text, a key with no payload, or a payload that isn't
    valid UTF-8 — since the input is just whatever _CF_EMAIL_RE happened to
    match in scraped HTML, not guaranteed-well-formed Cloudflare output."""
    try:
        data = bytes.fromhex(hex_str)
    except ValueError:
        return None
    if len(data) < 2:
        return None
    key = data[0]
    try:
        return bytes(b ^ key for b in data[1:]).decode("utf-8")
    except UnicodeDecodeError:
        return None


def extract_contact_email(html: str, allowed_prefixes: frozenset[str]) -> Optional[str]:
    """Return the first valid email in the page whose local-part matches the
    prefix allowlist. Candidates are ranked mailto: links first, then
    Cloudflare email-obfuscation spans (data-cfemail; see _decode_cf_email),
    then plain-text matches; asset filenames like 'sprite@2x.png' are
    ignored, as are addresses whose prefix is not on the allowlist (e.g.
    personal name.surname@ addresses)."""
    cf_candidates = [
        decoded
        for decoded in (_decode_cf_email(m) for m in _CF_EMAIL_RE.findall(html))
        if decoded is not None
    ]
    candidates = _MAILTO_RE.findall(html) + cf_candidates + _TEXT_EMAIL_RE.findall(html)
    for candidate in candidates:
        candidate = candidate.strip().rstrip(".")
        if _is_asset_filename(candidate):
            continue
        try:
            validated = validate_email(candidate, check_deliverability=False)
        except EmailNotValidError:
            continue
        local_part = validated.normalized.rsplit("@", 1)[0]
        if not _prefix_allowed(local_part, allowed_prefixes):
            continue
        return validated.normalized
    return None


def _extract_form_blocks(html: str) -> list[str]:
    """Return the substring of every <form>...</form> block in html, in
    document order. An unclosed <form> (no matching </form>) uses the rest
    of the document as its block rather than being skipped, since malformed
    markup shouldn't hide a real form that lacks a closing tag. Overlapping
    <form> opens found inside an already-consumed block (nested/malformed
    markup) are skipped so they aren't double-counted."""
    blocks: list[str] = []
    pos = 0
    for open_match in _FORM_OPEN_RE.finditer(html):
        if open_match.start() < pos:
            continue
        close_match = _FORM_CLOSE_RE.search(html, open_match.end())
        end = close_match.end() if close_match else len(html)
        blocks.append(html[open_match.start() : end])
        pos = end
    return blocks


def _form_block_has_signal(block: str) -> bool:
    """Whether a single <form>...</form> block satisfies the contact-form
    heuristic: a message-type field (_MESSAGE_FIELD_RE), or contact-related
    keyword text anywhere in the block (_CONTACT_FORM_KEYWORDS). A bare
    newsletter-signup form with only an <input type="email"> is accepted as
    a known false positive of this heuristic — see scrape_domain_contact."""
    if _MESSAGE_FIELD_RE.search(block):
        return True
    lowered = block.lower()
    return any(keyword in lowered for keyword in _CONTACT_FORM_KEYWORDS)


def has_contact_form(html: str) -> bool:
    """Whether html contains at least one <form> block matching the
    contact-form heuristic. A <form> with no supporting signal (e.g. a
    search box: <input type="text"> and a submit button, nothing else) does
    not count — filtering that out is the whole point of requiring a
    signal, scoped to within each form block."""
    return any(_form_block_has_signal(block) for block in _extract_form_blocks(html))


def scrape_domain_contact(
    host: str,
    slugs: list[str],
    rp: RobotFileParser,
    user_agent: str,
    delay: float,
    logger: logging.Logger,
    allowed_prefixes: frozenset[str],
    http_get: Callable[..., requests.Response] = requests.get,
    sleep_fn: Callable[[float], None] = time.sleep,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Try each candidate page for host in order; return (email_url, email,
    form_url). email_url/email are for the first page that yields an
    allowlisted email (or (None, None) if none do). form_url is the first
    candidate page — checked in this same fetch loop, against the same
    already-fetched HTML, no extra requests — found to contain a contact
    form (has_contact_form), or None if none do. Because the loop returns as
    soon as an email is found, pages after an early email hit are never
    checked for a form; the URL recorded for the form need not be the same
    page as the one recorded for the email."""
    form_url: Optional[str] = None
    for url in build_candidate_urls(host, slugs):
        if not rp.can_fetch(user_agent, url):
            logger.info("robots.txt disallows %s; skipping", url)
            continue
        sleep_fn(delay)
        logger.info("Fetching %s", url)
        html = fetch_page(
            url,
            user_agent,
            logger,
            http_get=http_get,
            resolve_host=resolve_host,
        )
        if html is None:
            continue
        if form_url is None and has_contact_form(html):
            logger.info("Found contact form at %s", url)
            form_url = url
        email = extract_contact_email(html, allowed_prefixes)
        if email:
            logger.info("Found contact email %s at %s", email, url)
            return url, email, form_url
        logger.info("Content at %s has no email; trying next page", url)
    return None, None, form_url


def _find_column(header_row: list, name: str) -> int:
    """1-based index of the named column in the workbook header."""
    for index, value in enumerate(header_row, start=1):
        if value == name:
            return index
    raise ValueError(f"Results workbook has no {name!r} column")


def _write_contact_cell(sheet, row: int, column: int, value: Optional[str]) -> None:
    cell = sheet.cell(row=row, column=column)
    if value is None:
        # Leave the cell empty rather than writing the literal text "None" —
        # a downstream re-read of the workbook can't otherwise distinguish
        # "nothing found" from a cell that legitimately contains that text.
        # The red fill is the "nothing found" signal instead.
        cell.value = None
        cell.fill = PatternFill(fill_type="solid", fgColor=schemas.NONE_FILL_COLOR)
    else:
        cell.value = value


@dataclass(frozen=True)
class ContactScrapeStats:
    processed: int
    found: int
    form_found: int


def populate_contact_columns(
    workbook,
    sheet,
    domain_col: int,
    url_col: int,
    email_col: int,
    form_col: int,
    xlsx_path: Path,
    slugs: list[str],
    user_agent: str,
    logger: logging.Logger,
    allowed_prefixes: frozenset[str],
    http_get: Callable[..., requests.Response] = requests.get,
    sleep_fn: Callable[[float], None] = time.sleep,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> ContactScrapeStats:
    """Scrape a contact url/email/form for each domain row in sheet and write it
    into url_col/email_col/form_col, caching robots.txt per host. Checkpoints
    the workbook to disk every schemas.CHECKPOINT_EVERY_ROWS processed rows
    (instead of every row, which rewrites the whole file each time — O(n^2)
    bytes written for N rows), plus a final save after the loop so nothing
    from the last partial batch is lost. Returns processed/found/form_found
    counts for the caller's summary log."""
    robots_cache: dict[str, tuple[RobotFileParser, float]] = {}
    processed = 0
    found = 0
    form_found = 0
    for row in range(2, sheet.max_row + 1):
        domain = sheet.cell(row=row, column=domain_col).value
        if domain is None or not str(domain).strip():
            continue
        host = str(domain).strip()

        if host not in robots_cache:
            rp = load_robots(host, user_agent, logger, http_get, sleep_fn, resolve_host)
            robots_cache[host] = (rp, resolve_crawl_delay(rp, user_agent))
        rp, delay = robots_cache[host]

        logger.info("Scraping contact for domain=%s", host)
        contact_url, contact_email, contact_form_url = scrape_domain_contact(
            host,
            slugs,
            rp,
            user_agent,
            delay,
            logger,
            allowed_prefixes,
            http_get,
            sleep_fn,
            resolve_host,
        )
        if contact_email is None:
            logger.info("No contact email found for domain=%s", host)
        if contact_form_url is None:
            logger.info("No contact form found for domain=%s", host)

        _write_contact_cell(sheet, row, url_col, contact_url)
        _write_contact_cell(sheet, row, email_col, contact_email)
        _write_contact_cell(sheet, row, form_col, contact_form_url)

        processed += 1
        if contact_email is not None:
            found += 1
        if contact_form_url is not None:
            form_found += 1
        if processed % schemas.CHECKPOINT_EVERY_ROWS == 0:
            workbook.save(str(xlsx_path))

    workbook.save(str(xlsx_path))  # final flush: last partial batch + full run
    return ContactScrapeStats(processed=processed, found=found, form_found=form_found)


def augment_workbook_with_contacts(
    xlsx_path: Path,
    slugs: list[str],
    user_agent: str,
    logger: logging.Logger,
    allowed_prefixes: frozenset[str],
    http_get: Callable[..., requests.Response] = requests.get,
    sleep_fn: Callable[[float], None] = time.sleep,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> Path:
    """Add contact_url / contact_email / formularz kontaktowy columns to the
    results workbook, one row per domain, colouring cells red where nothing
    was found. Checkpoints periodically (see populate_contact_columns) plus a
    final save, so a long run's progress survives an interruption without
    rewriting the whole file every row."""
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    header_row = [
        sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)
    ]
    domain_col = _find_column(header_row, "domain")
    url_col = len(header_row) + 1
    email_col = len(header_row) + 2
    form_col = len(header_row) + 3
    sheet.cell(row=1, column=url_col, value=schemas.CONTACT_URL_HEADER)
    sheet.cell(row=1, column=email_col, value=schemas.CONTACT_EMAIL_HEADER)
    sheet.cell(row=1, column=form_col, value=schemas.CONTACT_FORM_HEADER)

    stats = populate_contact_columns(
        workbook,
        sheet,
        domain_col,
        url_col,
        email_col,
        form_col,
        xlsx_path,
        slugs,
        user_agent,
        logger,
        allowed_prefixes,
        http_get,
        sleep_fn,
        resolve_host,
    )

    logger.info(
        "Contact scraping complete: %d domain(s) processed, %d with email, "
        "%d without (marked red), %d with contact form in %s",
        stats.processed,
        stats.found,
        stats.processed - stats.found,
        stats.form_found,
        xlsx_path,
    )
    return xlsx_path


def scrape_contacts(
    results_xlsx_path: Path,
    contact_email: str,
    about_us_csv_path: Path,
    email_prefixes_csv_path: Path,
    logger: logging.Logger,
    http_get: Callable[..., requests.Response] = requests.get,
    sleep_fn: Callable[[float], None] = time.sleep,
    resolve_host: Callable[[str], list] = _default_resolve_host,
) -> Path:
    """Entry point: read the about-us slug list and the email-prefix allowlist,
    then augment the results workbook with a contact url/email for each domain,
    keeping only emails whose local-part matches an allowlisted prefix."""
    slugs = read_about_us_slugs(about_us_csv_path)
    allowed_prefixes = read_email_prefixes(email_prefixes_csv_path)
    if not allowed_prefixes:
        logger.warning(
            "Email-prefix allowlist %s is empty; no contact emails will be kept",
            email_prefixes_csv_path,
        )
    user_agent = build_user_agent(contact_email)
    logger.info(
        "Starting contact scrape for %s using %d candidate slug(s) and %d "
        "allowlisted prefix(es); user_agent=%s",
        results_xlsx_path,
        len(slugs),
        len(allowed_prefixes),
        user_agent,
    )
    return augment_workbook_with_contacts(
        results_xlsx_path,
        slugs,
        user_agent,
        logger,
        allowed_prefixes,
        http_get=http_get,
        sleep_fn=sleep_fn,
        resolve_host=resolve_host,
    )
