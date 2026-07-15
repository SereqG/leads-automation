import logging
import socket
from urllib.robotparser import RobotFileParser

import pytest
import requests
from openpyxl import Workbook, load_workbook

from apps.enrichment import schemas, services, tasks
from core.exceptions import ValidationFailedError

LOGGER = logging.getLogger("test-enrichment")

# Allowlist used across email-extraction/scraping tests.
ALLOWED_PREFIXES = frozenset({"info", "hello", "kontakt", "office", "sales"})


def _resolve_public(host):
    """Fake resolver (mirrors socket.getaddrinfo's return shape) that always
    reports a public IP, so tests stay hermetic instead of doing a real DNS
    lookup via _is_public_host's default resolver."""
    return [(socket.AF_INET, socket.SOCK_STREAM, 6, "", ("93.184.216.34", 0))]


def _resolve_private(host):
    return [(socket.AF_INET, socket.SOCK_STREAM, 6, "", ("10.0.0.5", 0))]


class FakeResponse:
    def __init__(self, text="", status_code=200, headers=None):
        self.text = text
        self.status_code = status_code
        self.headers = headers if headers is not None else {"Content-Type": "text/html"}


# --- read_about_us_slugs ---


def test_read_about_us_slugs_dedupes_preserving_order(tmp_path):
    csv_path = tmp_path / "about.csv"
    csv_path.write_text(
        "page_name,url_slug\n"
        "O nas,/o-nas\n"
        "BOK,/bok\n"
        "Biuro,/bok\n"  # duplicate slug
        "FAQ,/faq\n"
        "Blank,\n"  # empty slug -> skipped
    )
    assert services.read_about_us_slugs(csv_path) == ["/o-nas", "/bok", "/faq"]


# --- read_email_prefixes ---


def test_read_email_prefixes_lowercases_strips_and_dedupes(tmp_path):
    csv_path = tmp_path / "email_prefixes.csv"
    csv_path.write_text(
        "prefix\n"
        "info\n"
        "  Office  \n"  # surrounding whitespace + mixed case
        "INFO\n"  # duplicate after lowercasing
        "\n"  # blank -> skipped
        "kontakt\n"
    )
    assert services.read_email_prefixes(csv_path) == frozenset(
        {"info", "office", "kontakt"}
    )


# --- build_candidate_urls ---


def test_build_candidate_urls_homepage_first():
    assert services.build_candidate_urls("example.com", ["/o-nas", "/kontakt"]) == [
        "https://example.com/",
        "https://example.com/o-nas",
        "https://example.com/kontakt",
    ]


# --- extract_contact_email ---


def test_extract_contact_email_prefers_mailto_over_plain_text():
    html = 'plain info@plain.com <a href="mailto:hello@shop.pl?subject=Hi">write</a>'
    assert services.extract_contact_email(html, ALLOWED_PREFIXES) == "hello@shop.pl"


def test_extract_contact_email_plain_text_normalized_and_trimmed():
    # The trailing sentence period is stripped and the address is normalised.
    assert (
        services.extract_contact_email(
            "Reach us at Info@Example.COM.", ALLOWED_PREFIXES
        )
        == "info@example.com"
    )


def test_extract_contact_email_rejects_asset_filenames():
    assert (
        services.extract_contact_email('<img src="sprite@2x.png">', ALLOWED_PREFIXES)
        is None
    )


def test_extract_contact_email_none_when_absent():
    assert (
        services.extract_contact_email("no email here, just words", ALLOWED_PREFIXES)
        is None
    )


def test_extract_contact_email_keeps_allowlisted_prefixes():
    # Exact prefix and prefix-plus-separator variants are all kept.
    assert (
        services.extract_contact_email("Write to info@company.pl", ALLOWED_PREFIXES)
        == "info@company.pl"
    )
    assert (
        services.extract_contact_email(
            "Write to info.sales@company.pl", ALLOWED_PREFIXES
        )
        == "info.sales@company.pl"
    )
    assert (
        services.extract_contact_email("Write to info-pl@company.pl", ALLOWED_PREFIXES)
        == "info-pl@company.pl"
    )


def test_extract_contact_email_drops_non_allowlisted_prefixes():
    # Personal addresses and prefixes without a separator boundary are dropped.
    assert (
        services.extract_contact_email(
            "Write to name.surname@company.pl", ALLOWED_PREFIXES
        )
        is None
    )
    assert (
        services.extract_contact_email(
            "Write to informacje@company.pl", ALLOWED_PREFIXES
        )
        is None
    )


def test_extract_contact_email_prefix_match_is_case_insensitive():
    assert (
        services.extract_contact_email("Write to INFO@Company.PL", ALLOWED_PREFIXES)
        == "info@company.pl"
    )


def test_extract_contact_email_skips_personal_and_returns_allowlisted():
    # A page with both a personal and a generic address keeps the generic one.
    html = "Jan: jan.kowalski@company.pl or general info@company.pl"
    assert services.extract_contact_email(html, ALLOWED_PREFIXES) == "info@company.pl"


def test_extract_contact_email_decodes_cloudflare_obfuscation():
    # Captured from tanisklepmedyczny.pl (Cloudflare Scrape Shield enabled):
    # no mailto: link and no plain-text address anywhere in the HTML, only
    # this hex+XOR-encoded span, decoding to biuro@tanisklepmedyczny.pl.
    html = (
        '<a href="/cdn-cgi/l/email-protection#">'
        '<span class="__cf_email__" '
        'data-cfemail="086a617d7a67487c6966617b63646d78656d6c716b726671267864">'
        "[email&#160;protected]</span></a>"
    )
    assert (
        services.extract_contact_email(html, frozenset({"biuro"}))
        == "biuro@tanisklepmedyczny.pl"
    )


def test_extract_contact_email_ignores_malformed_cf_obfuscation():
    # Odd-length hex can't be a key byte + payload; decode fails gracefully
    # and extraction falls through to the next candidate instead of raising.
    html = '<span data-cfemail="abc">bad</span> Write to kontakt@firma.pl'
    assert services.extract_contact_email(html, ALLOWED_PREFIXES) == "kontakt@firma.pl"


def test_extract_contact_email_prefers_mailto_over_cf_obfuscation():
    html = (
        '<a href="mailto:hello@shop.pl">write</a> '
        '<span data-cfemail="2a4145445e4b415e6a4c4358474b045a46">cf</span>'
    )
    assert services.extract_contact_email(html, ALLOWED_PREFIXES) == "hello@shop.pl"


def test_extract_contact_email_prefers_cf_obfuscation_over_plain_text():
    html = (
        "plain sales@plain.com "
        '<span data-cfemail="2a4145445e4b415e6a4c4358474b045a46">cf</span>'
    )
    assert services.extract_contact_email(html, ALLOWED_PREFIXES) == "kontakt@firma.pl"


# --- _is_public_host (SSRF guard) ---


def test_is_public_host_true_for_public_ip():
    assert services._is_public_host("example.com", resolve_host=_resolve_public) is True


def test_is_public_host_false_for_private_ip():
    resolve = lambda host: [(socket.AF_INET, 0, 0, "", ("10.0.0.5", 0))]
    assert services._is_public_host("internal.example.com", resolve_host=resolve) is False


def test_is_public_host_false_for_loopback_ip():
    resolve = lambda host: [(socket.AF_INET, 0, 0, "", ("127.0.0.1", 0))]
    assert services._is_public_host("localhost", resolve_host=resolve) is False


def test_is_public_host_false_for_link_local_metadata_ip():
    resolve = lambda host: [(socket.AF_INET, 0, 0, "", ("169.254.169.254", 0))]
    assert services._is_public_host("metadata.internal", resolve_host=resolve) is False


def test_is_public_host_false_on_dns_failure():
    def resolve(host):
        raise socket.gaierror("no such host")

    assert services._is_public_host("nonexistent.invalid", resolve_host=resolve) is False


# --- fetch_page ---


def test_fetch_page_returns_text_for_html():
    resp = FakeResponse(
        text="<html>ok</html>", headers={"Content-Type": "text/html; charset=utf-8"}
    )
    assert (
        services.fetch_page(
            "https://x.com/",
            "ua",
            LOGGER,
            http_get=lambda *a, **k: resp,
            resolve_host=_resolve_public,
        )
        == "<html>ok</html>"
    )


def test_fetch_page_none_on_non_200():
    resp = FakeResponse(status_code=404)
    assert (
        services.fetch_page(
            "https://x.com/x",
            "ua",
            LOGGER,
            http_get=lambda *a, **k: resp,
            resolve_host=_resolve_public,
        )
        is None
    )


def test_fetch_page_none_on_non_html_content():
    resp = FakeResponse(text="%PDF", headers={"Content-Type": "application/pdf"})
    assert (
        services.fetch_page(
            "https://x.com/f",
            "ua",
            LOGGER,
            http_get=lambda *a, **k: resp,
            resolve_host=_resolve_public,
        )
        is None
    )


def test_fetch_page_none_on_non_allowlisted_text_content():
    # Tightened allowlist: text/csv, text/plain etc. no longer pass just
    # because "text" is a substring of the Content-Type.
    resp = FakeResponse(text="a,b,c", headers={"Content-Type": "text/csv"})
    assert (
        services.fetch_page(
            "https://x.com/data.csv",
            "ua",
            LOGGER,
            http_get=lambda *a, **k: resp,
            resolve_host=_resolve_public,
        )
        is None
    )


def test_fetch_page_none_when_body_exceeds_size_cap():
    resp = FakeResponse(
        text="a" * (schemas.MAX_PAGE_BYTES + 1),
        headers={"Content-Type": "text/html"},
    )
    assert (
        services.fetch_page(
            "https://x.com/huge",
            "ua",
            LOGGER,
            http_get=lambda *a, **k: resp,
            resolve_host=_resolve_public,
        )
        is None
    )


def test_fetch_page_none_on_request_exception():
    def boom(*args, **kwargs):
        raise requests.ConnectionError("boom")

    assert (
        services.fetch_page(
            "https://x.com/", "ua", LOGGER, http_get=boom, resolve_host=_resolve_public
        )
        is None
    )


def test_fetch_page_none_when_host_not_public():
    def http_get(*args, **kwargs):
        raise AssertionError("http_get should not be called for a non-public host")

    assert (
        services.fetch_page(
            "https://x.com/",
            "ua",
            LOGGER,
            http_get=http_get,
            resolve_host=_resolve_private,
        )
        is None
    )


@pytest.mark.parametrize("status_code", [301, 302, 303, 307, 308])
def test_fetch_page_redirect_is_not_followed(status_code):
    # A redirected candidate page (e.g. www.orteo.pl sending every candidate
    # page to a captcha interstitial) is treated as unavailable rather than
    # chased, so the caller moves straight on to the next candidate page.
    fetched = []

    def http_get(url, headers, timeout, **kwargs):
        fetched.append(url)
        return FakeResponse(
            status_code=status_code, headers={"Location": "/captcha.php"}
        )

    assert (
        services.fetch_page(
            "https://x.com/go",
            "ua",
            LOGGER,
            http_get=http_get,
            resolve_host=_resolve_public,
        )
        is None
    )
    # The redirect target must never actually be requested.
    assert fetched == ["https://x.com/go"]


# --- load_robots ---


def test_load_robots_parses_disallow_rules():
    resp = FakeResponse(
        text="User-agent: *\nDisallow: /private\n",
        headers={"Content-Type": "text/plain"},
    )
    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=lambda *a, **k: resp,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/private/x") is False
    assert rp.can_fetch("ua", "https://example.com/o-nas") is True


def test_load_robots_allows_all_on_404():
    resp = FakeResponse(status_code=404)
    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=lambda *a, **k: resp,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is True


def test_load_robots_disallows_all_on_403():
    resp = FakeResponse(status_code=403)
    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=lambda *a, **k: resp,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is False


def test_load_robots_disallows_all_on_request_exception():
    def boom(*args, **kwargs):
        raise requests.ConnectionError("boom")

    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=boom,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is False


def test_load_robots_disallows_all_on_5xx():
    resp = FakeResponse(status_code=503)
    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=lambda *a, **k: resp,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is False


def test_load_robots_disallows_all_when_host_not_public():
    def http_get(*args, **kwargs):
        raise AssertionError("http_get should not be called for a non-public host")

    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_private,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is False


def test_load_robots_follows_same_host_redirect():
    fetched = []

    def http_get(url, headers, timeout, **kwargs):
        fetched.append(url)
        if url == "https://example.com/robots.txt":
            return FakeResponse(
                status_code=301,
                headers={"Location": "https://example.com/new-robots.txt"},
            )
        return FakeResponse(text="User-agent: *\nDisallow: /private\n")

    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert fetched == [
        "https://example.com/robots.txt",
        "https://example.com/new-robots.txt",
    ]
    assert rp.can_fetch("ua", "https://example.com/private/x") is False
    assert rp.can_fetch("ua", "https://example.com/o-nas") is True


def test_load_robots_disallows_all_on_cross_host_redirect():
    resp = FakeResponse(
        status_code=302, headers={"Location": "https://evil.example.net/robots.txt"}
    )
    rp = services.load_robots(
        "example.com",
        "ua",
        LOGGER,
        http_get=lambda *a, **k: resp,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )
    assert rp.can_fetch("ua", "https://example.com/anything") is False


# --- has_contact_form ---


def test_has_contact_form_true_with_textarea():
    assert services.has_contact_form("<form><textarea></textarea></form>") is True


def test_has_contact_form_true_with_email_input():
    html = '<form><input type="email" name="reply_to"></form>'
    assert services.has_contact_form(html) is True


def test_has_contact_form_true_with_keyword_only_text():
    html = '<form><p>Napisz do nas</p><input type="text" name="name"></form>'
    assert services.has_contact_form(html) is True


def test_has_contact_form_false_for_signal_less_search_box():
    html = (
        '<form role="search"><input type="text" name="q"><button>Search</button></form>'
    )
    assert services.has_contact_form(html) is False


def test_has_contact_form_false_when_no_form_tag_present():
    # A signal with no enclosing <form> at all doesn't count.
    assert services.has_contact_form("<textarea>free text, no form</textarea>") is False
    assert services.has_contact_form("<p>Kontakt: biuro@example.com</p>") is False


def test_has_contact_form_signal_scoped_to_form_block():
    # The keyword sits outside the <form>; the form itself has no signal.
    html = '<p>Kontakt</p><form><input type="text"></form>'
    assert services.has_contact_form(html) is False


def test_has_contact_form_multiple_blocks_only_one_needs_signal():
    html = (
        '<form role="search"><input type="text"></form>'
        "<form><textarea></textarea></form>"
    )
    assert services.has_contact_form(html) is True


def test_has_contact_form_case_insensitive():
    assert services.has_contact_form("<FORM><TEXTAREA></TEXTAREA></FORM>") is True
    assert services.has_contact_form('<FORM><input TYPE="EMAIL"></FORM>') is True
    assert (
        services.has_contact_form('<FORM><input type="text">CONTACT US</FORM>') is True
    )


def test_has_contact_form_true_for_bare_newsletter_signup_known_false_positive():
    # Accepted tradeoff of this heuristic (design decision, not a bug): a
    # bare newsletter-signup form with just an email field still counts.
    html = '<form><input type="email" placeholder="Your email"><button>Subscribe</button></form>'
    assert services.has_contact_form(html) is True


def test_has_contact_form_unclosed_form_tag_still_scanned():
    html = "<div><form><p>get in touch</p>"
    assert services.has_contact_form(html) is True


def test_has_contact_form_false_for_empty_html():
    assert services.has_contact_form("") is False


# --- scrape_domain_contact ---


def test_scrape_domain_contact_skips_disallowed_and_finds_next():
    rp = RobotFileParser()
    rp.parse(["User-agent: *", "Disallow: /o-nas"])

    fetched = []

    def http_get(url, headers, timeout, **kwargs):
        fetched.append(url)
        if url == "https://example.com/":
            return FakeResponse(text="homepage, no email")
        if url == "https://example.com/kontakt":
            return FakeResponse(text='<a href="mailto:kontakt@example.com">mail</a>')
        return FakeResponse(status_code=404)

    url, email, form_url = services.scrape_domain_contact(
        "example.com",
        ["/o-nas", "/kontakt"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert (url, email, form_url) == (
        "https://example.com/kontakt",
        "kontakt@example.com",
        None,
    )
    # /o-nas is disallowed by robots.txt and must never be requested.
    assert "https://example.com/o-nas" not in fetched
    assert fetched == ["https://example.com/", "https://example.com/kontakt"]


def test_scrape_domain_contact_returns_none_when_nothing_found():
    rp = RobotFileParser()
    rp.allow_all = True

    def http_get(url, headers, timeout, **kwargs):
        return FakeResponse(status_code=404)

    assert services.scrape_domain_contact(
        "example.com",
        ["/o-nas"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    ) == (None, None, None)


def test_scrape_domain_contact_skips_personal_email_and_keeps_looking():
    rp = RobotFileParser()
    rp.allow_all = True

    def http_get(url, headers, timeout, **kwargs):
        if url == "https://example.com/":
            # Only a personal address on the homepage -> filtered out.
            return FakeResponse(text="Jan Kowalski: jan.kowalski@example.com")
        if url == "https://example.com/kontakt":
            return FakeResponse(text='<a href="mailto:kontakt@example.com">mail</a>')
        return FakeResponse(status_code=404)

    url, email, form_url = services.scrape_domain_contact(
        "example.com",
        ["/kontakt"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert (url, email, form_url) == (
        "https://example.com/kontakt",
        "kontakt@example.com",
        None,
    )


def test_scrape_domain_contact_returns_form_url_from_earlier_page_than_email():
    rp = RobotFileParser()
    rp.allow_all = True

    def http_get(url, headers, timeout, **kwargs):
        if url == "https://example.com/":
            # Form but no email.
            return FakeResponse(text="<form><textarea></textarea></form>")
        if url == "https://example.com/kontakt":
            return FakeResponse(text='<a href="mailto:kontakt@example.com">mail</a>')
        return FakeResponse(status_code=404)

    result = services.scrape_domain_contact(
        "example.com",
        ["/kontakt"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert result == (
        "https://example.com/kontakt",
        "kontakt@example.com",
        "https://example.com/",
    )


def test_scrape_domain_contact_email_and_form_on_same_page():
    rp = RobotFileParser()
    rp.allow_all = True

    def http_get(url, headers, timeout, **kwargs):
        if url == "https://example.com/":
            return FakeResponse(
                text="<form><textarea></textarea></form>"
                '<a href="mailto:kontakt@example.com">mail</a>'
            )
        return FakeResponse(status_code=404)

    result = services.scrape_domain_contact(
        "example.com",
        [],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert result == (
        "https://example.com/",
        "kontakt@example.com",
        "https://example.com/",
    )


def test_scrape_domain_contact_stops_form_search_after_early_email_hit():
    # An email found on the homepage stops the loop there, so /kontakt is
    # never fetched — even though it would have had a form. Pins down the
    # "reuse existing fetches, no extra requests" design decision.
    rp = RobotFileParser()
    rp.allow_all = True

    fetched = []

    def http_get(url, headers, timeout, **kwargs):
        fetched.append(url)
        if url == "https://example.com/":
            return FakeResponse(text='<a href="mailto:kontakt@example.com">mail</a>')
        if url == "https://example.com/kontakt":
            return FakeResponse(text="<form><textarea></textarea></form>")
        return FakeResponse(status_code=404)

    result = services.scrape_domain_contact(
        "example.com",
        ["/kontakt"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert result == ("https://example.com/", "kontakt@example.com", None)
    assert fetched == ["https://example.com/"]


def test_scrape_domain_contact_records_first_form_when_multiple_pages_have_forms():
    rp = RobotFileParser()
    rp.allow_all = True

    def http_get(url, headers, timeout, **kwargs):
        if url == "https://example.com/":
            return FakeResponse(text="<form><textarea></textarea></form>")
        if url == "https://example.com/kontakt":
            return FakeResponse(text="<form><textarea></textarea></form>")
        return FakeResponse(status_code=404)

    result = services.scrape_domain_contact(
        "example.com",
        ["/kontakt"],
        rp,
        "ua",
        3,
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert result == (None, None, "https://example.com/")


# --- populate_contact_columns ---


def test_populate_contact_columns_writes_cells_and_returns_stats(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "domain"])
    sheet.append([1, "found.com"])
    sheet.append([2, "none.com"])
    sheet.append([3, ""])  # blank domain -> skipped entirely
    path = tmp_path / "results.xlsx"
    workbook.save(path)

    def http_get(url, headers, timeout, **kwargs):
        if url.endswith("/robots.txt"):
            return FakeResponse(status_code=404)  # allow all
        if url == "https://found.com/":
            return FakeResponse(text='<a href="mailto:info@found.com">contact</a>')
        return FakeResponse(status_code=404)  # none.com yields nothing

    stats = services.populate_contact_columns(
        workbook,
        sheet,
        domain_col=2,
        url_col=3,
        email_col=4,
        form_col=5,
        xlsx_path=path,
        slugs=["/kontakt"],
        user_agent="ua",
        logger=LOGGER,
        allowed_prefixes=ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert stats == services.ContactScrapeStats(processed=2, found=1, form_found=0)

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == (1, "found.com", "https://found.com/", "info@found.com", None)
    assert rows[1] == (2, "none.com", None, None, None)

    # Row 4 (blank domain) is untouched — no url/email/form cells written.
    assert sheet.cell(row=4, column=3).value is None
    assert sheet.cell(row=4, column=4).value is None
    assert sheet.cell(row=4, column=5).value is None

    # Only 2 rows are processed here, well under CHECKPOINT_EVERY_ROWS, so no
    # mid-loop checkpoint fires — but the final save after the loop always
    # runs, so the result is still durably persisted by the time the call
    # returns. Confirmed by re-reading from disk.
    persisted = list(load_workbook(path).active.iter_rows(min_row=2, values_only=True))
    assert persisted[0] == (
        1,
        "found.com",
        "https://found.com/",
        "info@found.com",
        None,
    )


def test_populate_contact_columns_writes_form_url_cell_and_counts_form_found(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "domain"])
    sheet.append([1, "formonly.com"])
    path = tmp_path / "results.xlsx"
    workbook.save(path)

    def http_get(url, headers, timeout, **kwargs):
        if url.endswith("/robots.txt"):
            return FakeResponse(status_code=404)  # allow all
        if url == "https://formonly.com/":
            return FakeResponse(text="<form><textarea></textarea></form>")
        return FakeResponse(status_code=404)

    stats = services.populate_contact_columns(
        workbook,
        sheet,
        domain_col=2,
        url_col=3,
        email_col=4,
        form_col=5,
        xlsx_path=path,
        slugs=[],
        user_agent="ua",
        logger=LOGGER,
        allowed_prefixes=ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert stats == services.ContactScrapeStats(processed=1, found=0, form_found=1)

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == (1, "formonly.com", None, None, "https://formonly.com/")

    # url/email cells: empty and red-filled ("not found"); form cell: no fill.
    for column in (3, 4):
        cell = sheet.cell(row=2, column=column)
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb == "FFFF0000"
    assert sheet.cell(row=2, column=5).fill.fill_type is None


def test_populate_contact_columns_checkpoints_every_k_rows_not_every_row(tmp_path):
    # Regression test for the O(n^2) full-workbook rewrite: workbook.save
    # must be called on a checkpoint cadence (schemas.CHECKPOINT_EVERY_ROWS),
    # not once per processed row.
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "domain"])
    n_domains = 30
    for i in range(1, n_domains + 1):
        sheet.append([i, f"none{i}.com"])
    path = tmp_path / "results.xlsx"
    workbook.save(path)

    save_calls = []
    original_save = workbook.save

    def counting_save(*args, **kwargs):
        save_calls.append(1)
        return original_save(*args, **kwargs)

    workbook.save = counting_save

    def http_get(url, headers, timeout, **kwargs):
        return FakeResponse(status_code=404)  # robots allow-all; no email anywhere

    services.populate_contact_columns(
        workbook,
        sheet,
        domain_col=2,
        url_col=3,
        email_col=4,
        form_col=5,
        xlsx_path=path,
        slugs=[],
        user_agent="ua",
        logger=LOGGER,
        allowed_prefixes=ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=lambda s: None,
        resolve_host=_resolve_public,
    )

    assert schemas.CHECKPOINT_EVERY_ROWS == 25
    # 30 rows at CHECKPOINT_EVERY_ROWS=25: one mid-loop checkpoint (at row
    # 25) plus the final save after the loop = 2 total, not 30.
    assert len(save_calls) == 2


# --- augment_workbook_with_contacts (end to end) ---


def _make_results_workbook(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "domain", "query"])
    sheet.append([1, "found.com", "q1"])
    sheet.append([2, "none.com", "q2"])
    path = tmp_path / "results.xlsx"
    workbook.save(path)
    return path


def test_augment_workbook_adds_columns_and_marks_red(tmp_path):
    path = _make_results_workbook(tmp_path)

    def http_get(url, headers, timeout, **kwargs):
        if url.endswith("/robots.txt"):
            return FakeResponse(status_code=404)  # allow all
        if url == "https://found.com/":
            return FakeResponse(
                text='<a href="mailto:info@found.com">contact</a>'
                "<form><textarea></textarea></form>"
            )
        return FakeResponse(status_code=404)  # none.com yields nothing

    sleeps = []
    services.augment_workbook_with_contacts(
        path,
        ["/kontakt"],
        "ua",
        LOGGER,
        ALLOWED_PREFIXES,
        http_get=http_get,
        sleep_fn=sleeps.append,
        resolve_host=_resolve_public,
    )

    sheet = load_workbook(path).active
    header = [cell.value for cell in sheet[1]]
    assert header == [
        "ID",
        "domain",
        "query",
        "contact_url",
        "contact_email",
        "formularz kontaktowy",
    ]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == (
        1,
        "found.com",
        "q1",
        "https://found.com/",
        "info@found.com",
        "https://found.com/",
    )
    assert rows[1] == (2, "none.com", "q2", None, None, None)

    # found.com cells: no red fill
    assert sheet.cell(row=2, column=4).fill.fill_type is None
    assert sheet.cell(row=2, column=5).fill.fill_type is None
    assert sheet.cell(row=2, column=6).fill.fill_type is None
    # none.com cells: red fill
    for column in (4, 5, 6):
        cell = sheet.cell(row=3, column=column)
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb == "FFFF0000"

    # A polite delay is applied before every request (robots + each page).
    # Still 5, not 6+, even though form-checking is now involved — form
    # detection reuses the HTML already fetched for the email search and
    # issues no extra requests (see scrape_domain_contact).
    assert len(sleeps) == 5
    assert all(delay == 3 for delay in sleeps)


# --- validate_scrape_inputs_for_chain ---


@pytest.fixture
def chain_valid_env(tmp_path, monkeypatch):
    about_csv = tmp_path / "about-us-urls.csv"
    about_csv.write_text("page_name,url_slug\nO nas,/o-nas\n")
    prefixes_csv = tmp_path / "email_prefixes_only.csv"
    prefixes_csv.write_text("prefix\ninfo\nkontakt\n")
    logs_dir = tmp_path / "logs"

    monkeypatch.setattr(schemas, "ABOUT_US_CSV_PATH", about_csv)
    monkeypatch.setattr(schemas, "EMAIL_PREFIXES_CSV_PATH", prefixes_csv)
    monkeypatch.setattr(schemas, "DEFAULT_LOG_DIR", logs_dir)

    return {
        "about_csv": about_csv,
        "prefixes_csv": prefixes_csv,
        "logs_dir": logs_dir,
    }


def test_validate_scrape_inputs_for_chain_passes(chain_valid_env, tmp_path):
    config = services.validate_scrape_inputs_for_chain(
        contact_email="test@example.com", log_file=tmp_path / "custom.log"
    )
    assert config.contact_email == "test@example.com"
    assert config.about_us_csv_path == chain_valid_env["about_csv"]
    assert config.email_prefixes_csv_path == chain_valid_env["prefixes_csv"]


def test_validate_scrape_inputs_for_chain_invalid_email_raises(
    chain_valid_env, tmp_path
):
    with pytest.raises(ValidationFailedError) as exc_info:
        services.validate_scrape_inputs_for_chain(
            contact_email="not-an-email", log_file=tmp_path / "custom.log"
        )
    assert any("email" in message.lower() for message in exc_info.value.errors)


def test_validate_scrape_inputs_for_chain_missing_about_csv_raises(
    chain_valid_env, tmp_path, monkeypatch
):
    monkeypatch.setattr(schemas, "ABOUT_US_CSV_PATH", tmp_path / "missing.csv")
    with pytest.raises(ValidationFailedError) as exc_info:
        services.validate_scrape_inputs_for_chain(
            contact_email="test@example.com", log_file=tmp_path / "custom.log"
        )
    assert any(
        "about-us-urls.csv not found" in message for message in exc_info.value.errors
    )


def test_validate_scrape_inputs_for_chain_missing_prefixes_csv_raises(
    chain_valid_env, tmp_path, monkeypatch
):
    monkeypatch.setattr(schemas, "EMAIL_PREFIXES_CSV_PATH", tmp_path / "missing.csv")
    with pytest.raises(ValidationFailedError) as exc_info:
        services.validate_scrape_inputs_for_chain(
            contact_email="test@example.com", log_file=tmp_path / "custom.log"
        )
    assert any(
        "email_prefixes_only.csv not found" in message
        for message in exc_info.value.errors
    )


def test_validate_scrape_inputs_for_chain_has_no_xlsx_field(chain_valid_env, tmp_path):
    config = services.validate_scrape_inputs_for_chain(
        contact_email="test@example.com", log_file=tmp_path / "custom.log"
    )
    assert not hasattr(config, "results_xlsx_path")


# --- Celery task ---


def test_scrape_contacts_task_calls_service(tmp_path, monkeypatch):
    captured = {}
    expected_dest = tmp_path / "results.xlsx"

    def fake_scrape_contacts(**kwargs):
        captured.update(kwargs)
        return expected_dest

    monkeypatch.setattr(tasks.services, "scrape_contacts", fake_scrape_contacts)

    log_file = tmp_path / "run.log"
    about_csv = tmp_path / "about.csv"
    prefixes_csv = tmp_path / "email_prefixes.csv"

    result = tasks.scrape_contacts_task(
        str(expected_dest),
        "me@example.com",
        str(log_file),
        str(about_csv),
        str(prefixes_csv),
    )

    assert result == str(expected_dest)
    assert captured["results_xlsx_path"] == expected_dest
    assert captured["contact_email"] == "me@example.com"
    assert captured["about_us_csv_path"] == about_csv
    assert captured["email_prefixes_csv_path"] == prefixes_csv
