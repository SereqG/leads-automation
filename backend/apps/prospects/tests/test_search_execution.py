import logging
import re
from types import SimpleNamespace

import pytest
import requests
from openpyxl import load_workbook

from apps.prospects import schemas, services, tasks


def _write_blacklist(tmp_path, lines=()):
    path = tmp_path / "blacklist.txt"
    content = "# test blacklist\n" + "".join(f"{line}\n" for line in lines)
    path.write_text(content)
    return path


class FakeResponse:
    def __init__(self, json_data, status_code=200):
        self._json_data = json_data
        self.status_code = status_code

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.HTTPError(f"status {self.status_code}")

    def json(self):
        return self._json_data


# --- compute_pagination_plan ---


@pytest.mark.parametrize(
    "per_query, expected",
    [
        (5, [(20, 0)]),
        (20, [(20, 0)]),
        (45, [(20, 0), (20, 1), (20, 2)]),
        (200, [(20, i) for i in range(10)]),
        (201, [(20, i) for i in range(10)]),
        (500, [(20, i) for i in range(10)]),
    ],
)
def test_compute_pagination_plan(per_query, expected):
    # count stays fixed at BRAVE_MAX_COUNT on every page (Brave's offset is
    # measured in pages of `count`, so it can't shrink on the last page);
    # truncation down to per_query happens in fetch_brave_results instead.
    assert services.compute_pagination_plan(per_query) == expected


# --- read_queries ---


def test_read_queries_filters_blank_rows(tmp_path):
    csv_path = tmp_path / "queries.csv"
    csv_path.write_text(
        "google_search_query\nsite:.pl sklep\n\nsite:.pl kup online\n\n\n"
    )
    assert services.read_queries(csv_path) == [
        "site:.pl sklep",
        "site:.pl kup online",
    ]


# --- fetch_brave_results ---


def test_fetch_brave_results_paginates_across_offsets():
    calls = []

    def fake_http_get(url, headers, params, timeout):
        calls.append(params)
        offset = params["offset"]
        count = params["count"]
        # Real Brave semantics: count is constant across pages (offset skips
        # offset * count results), so every page returns a full page here.
        results = [{"url": f"https://example.com/{offset}-{i}"} for i in range(count)]
        return FakeResponse({"web": {"results": results}})

    sleeps = []
    urls = services.fetch_brave_results(
        "plumbers",
        45,
        "key",
        logging.getLogger("test-fetch-paginate"),
        http_get=fake_http_get,
        sleep_fn=sleeps.append,
    )

    assert [c["offset"] for c in calls] == [0, 1, 2]
    assert [c["count"] for c in calls] == [20, 20, 20]
    # 3 pages of 20 = 60 fetched, truncated down to per_query=45.
    assert len(urls) == 45
    assert sleeps == [5, 5, 5]


def test_fetch_brave_results_stops_early_on_short_page():
    calls = []

    def fake_http_get(url, headers, params, timeout):
        calls.append(params)
        return FakeResponse({"web": {"results": [{"url": "https://example.com/1"}]}})

    urls = services.fetch_brave_results(
        "plumbers",
        45,
        "key",
        logging.getLogger("test-fetch-short"),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    assert len(calls) == 1
    assert urls == ["https://example.com/1"]


def test_fetch_brave_results_stops_query_on_request_exception(caplog):
    def fake_http_get(url, headers, params, timeout):
        raise requests.ConnectionError("boom")

    sleeps = []
    logger = logging.getLogger("test-fetch-error")
    with caplog.at_level(logging.ERROR, logger="test-fetch-error"):
        urls = services.fetch_brave_results(
            "plumbers",
            20,
            "key",
            logger,
            http_get=fake_http_get,
            sleep_fn=sleeps.append,
        )

    assert urls == []
    assert sleeps == [5]
    assert any("Brave request/parse failed" in r.message for r in caplog.records)


def test_fetch_brave_results_stops_query_on_invalid_json(caplog):
    class BadJsonResponse:
        def raise_for_status(self):
            pass

        def json(self):
            raise ValueError("Expecting value")

    def fake_http_get(url, headers, params, timeout):
        return BadJsonResponse()

    sleeps = []
    logger = logging.getLogger("test-fetch-badjson")
    with caplog.at_level(logging.ERROR, logger="test-fetch-badjson"):
        urls = services.fetch_brave_results(
            "plumbers",
            20,
            "key",
            logger,
            http_get=fake_http_get,
            sleep_fn=sleeps.append,
        )

    assert urls == []
    assert sleeps == [5]
    assert any("Brave request/parse failed" in r.message for r in caplog.records)


def test_fetch_brave_results_warns_when_per_query_exceeds_cap(caplog):
    def fake_http_get(url, headers, params, timeout):
        count = params["count"]
        results = [
            {"url": f"https://example.com/{params['offset']}-{i}"} for i in range(count)
        ]
        return FakeResponse({"web": {"results": results}})

    logger = logging.getLogger("test-fetch-cap-warning")
    with caplog.at_level(logging.WARNING, logger="test-fetch-cap-warning"):
        urls = services.fetch_brave_results(
            "plumbers",
            250,
            "key",
            logger,
            http_get=fake_http_get,
            sleep_fn=lambda seconds: None,
        )

    assert len(urls) == 200
    assert any("exceeds Brave's max" in r.message for r in caplog.records)


# --- extract_domain ---


@pytest.mark.parametrize(
    "url, expected",
    [
        ("https://Example.com/path?q=1", "example.com"),
        ("https://shop.example.com", "shop.example.com"),
        ("https://www.example.com", "www.example.com"),
        ("http://example.com:8080/", "example.com"),
        ("not a url", None),
        ("", None),
    ],
)
def test_extract_domain(url, expected):
    assert services.extract_domain(url) == expected


# --- load_blacklist ---


def test_load_blacklist_ignores_comments_blank_lines_and_lowercases(tmp_path):
    path = tmp_path / "blacklist.txt"
    path.write_text(
        "# === section header ===\n"
        "Allegro.PL\n"
        "\n"
        "  olx.pl  \n"
        "# another comment\n"
        "AMAZON.PL\n"
    )

    assert services.load_blacklist(path) == {"allegro.pl", "olx.pl", "amazon.pl"}


# --- is_blacklisted ---


@pytest.mark.parametrize(
    "domain, blacklist, expected",
    [
        ("allegro.pl", {"allegro.pl"}, True),
        ("shop.amazon.pl", {"amazon.pl"}, True),
        ("www.allegro.pl", {"allegro.pl"}, True),
        ("myallegro.pl", {"allegro.pl"}, False),
        ("allegro.pl.evil.com", {"allegro.pl"}, False),
        ("example.com", {"allegro.pl", "olx.pl"}, False),
    ],
)
def test_is_blacklisted(domain, blacklist, expected):
    assert services.is_blacklisted(domain, blacklist) is expected


# --- filter_blacklisted_rows ---


def test_filter_blacklisted_rows_drops_matches_and_renumbers_ids():
    rows = [
        services.SearchResultRow(id=1, domain="allegro.pl", query="q1"),
        services.SearchResultRow(id=2, domain="a.com", query="q1"),
        services.SearchResultRow(id=3, domain="shop.allegro.pl", query="q2"),
        services.SearchResultRow(id=4, domain="b.com", query="q2"),
    ]

    result = services.filter_blacklisted_rows(rows, {"allegro.pl"})

    assert result == [
        services.SearchResultRow(id=1, domain="a.com", query="q1"),
        services.SearchResultRow(id=2, domain="b.com", query="q2"),
    ]


# --- collect_search_rows ---


def test_collect_search_rows_collects_across_queries():
    def fake_http_get(url, headers, params, timeout):
        query = params["q"]
        return FakeResponse(
            {"web": {"results": [{"url": f"https://{query}.example.com/page"}]}}
        )

    result = services.collect_search_rows(
        ["plumbers", "electricians"],
        per_query=1,
        api_key="key",
        logger=logging.getLogger("test-collect-basic"),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    assert result.rows == [
        services.SearchResultRow(id=1, domain="plumbers.example.com", query="plumbers"),
        services.SearchResultRow(
            id=2, domain="electricians.example.com", query="electricians"
        ),
    ]
    assert result.total_urls == 2
    assert result.unparseable_count == 0


def test_collect_search_rows_dedupes_domains_across_queries():
    def fake_http_get(url, headers, params, timeout):
        query = params["q"]
        results = [
            {"url": "https://shared.example.com/a"},
            {"url": f"https://{query}-only.example.com"},
        ]
        return FakeResponse({"web": {"results": results}})

    result = services.collect_search_rows(
        ["plumbers", "electricians"],
        per_query=2,
        api_key="key",
        logger=logging.getLogger("test-collect-dedup"),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    # The shared domain is only kept once, under the first query that returned it.
    assert result.rows == [
        services.SearchResultRow(id=1, domain="shared.example.com", query="plumbers"),
        services.SearchResultRow(
            id=2, domain="plumbers-only.example.com", query="plumbers"
        ),
        services.SearchResultRow(
            id=3, domain="electricians-only.example.com", query="electricians"
        ),
    ]
    assert result.total_urls == 4


def test_collect_search_rows_counts_unparseable_urls(caplog):
    def fake_http_get(url, headers, params, timeout):
        results = [{"url": "not a url"}, {"url": "https://kept.example.com"}]
        return FakeResponse({"web": {"results": results}})

    logger = logging.getLogger("test-collect-unparseable")
    with caplog.at_level(logging.WARNING, logger="test-collect-unparseable"):
        result = services.collect_search_rows(
            ["shops"],
            per_query=2,
            api_key="key",
            logger=logger,
            http_get=fake_http_get,
            sleep_fn=lambda seconds: None,
        )

    assert result.rows == [
        services.SearchResultRow(id=1, domain="kept.example.com", query="shops")
    ]
    assert result.total_urls == 2
    assert result.unparseable_count == 1
    assert any("Could not extract a domain" in r.message for r in caplog.records)


# --- write_results_xlsx ---


def test_write_results_xlsx_round_trip(tmp_path):
    rows = [
        services.SearchResultRow(id=1, domain="a.com", query="q1"),
        services.SearchResultRow(id=2, domain="b.com", query="q1"),
    ]
    dest_path = services.write_results_xlsx(rows, tmp_path)

    assert dest_path.parent == tmp_path / "results"
    assert dest_path.exists()
    assert re.fullmatch(r"\d{8}_\d{6}\.xlsx", dest_path.name)

    workbook = load_workbook(dest_path)
    values = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    assert values == [
        ("ID", "domain", "query"),
        (1, "a.com", "q1"),
        (2, "b.com", "q1"),
    ]


# --- run_prospect_search ---


def test_run_prospect_search_end_to_end(tmp_path):
    queries_csv = tmp_path / "queries.csv"
    queries_csv.write_text("google_search_query\nplumbers\nelectricians\n")

    def fake_http_get(url, headers, params, timeout):
        query = params["q"]
        return FakeResponse(
            {"web": {"results": [{"url": f"https://{query}.example.com/page"}]}}
        )

    output_dir = tmp_path / "data"
    output_dir.mkdir()

    dest_path = services.run_prospect_search(
        queries_csv,
        per_query=1,
        api_key="key",
        logger=logging.getLogger("test-run-e2e"),
        output_dir=output_dir,
        blacklist_path=_write_blacklist(tmp_path),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    rows = list(load_workbook(dest_path).active.iter_rows(values_only=True))
    assert rows[0] == ("ID", "domain", "query")
    assert rows[1] == (1, "plumbers.example.com", "plumbers")
    assert rows[2] == (2, "electricians.example.com", "electricians")


def test_run_prospect_search_deduplicates_domains_across_queries(tmp_path):
    queries_csv = tmp_path / "queries.csv"
    queries_csv.write_text("google_search_query\nplumbers\nelectricians\n")

    def fake_http_get(url, headers, params, timeout):
        query = params["q"]
        results = [
            {"url": "https://shared.example.com/a"},
            {"url": f"https://{query}-only.example.com"},
        ]
        return FakeResponse({"web": {"results": results}})

    output_dir = tmp_path / "data"
    output_dir.mkdir()

    dest_path = services.run_prospect_search(
        queries_csv,
        per_query=2,
        api_key="key",
        logger=logging.getLogger("test-run-dedup"),
        output_dir=output_dir,
        blacklist_path=_write_blacklist(tmp_path),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    rows = list(load_workbook(dest_path).active.iter_rows(values_only=True))
    # The shared domain is only kept once, under the first query that returned it.
    assert rows == [
        ("ID", "domain", "query"),
        (1, "shared.example.com", "plumbers"),
        (2, "plumbers-only.example.com", "plumbers"),
        (3, "electricians-only.example.com", "electricians"),
    ]


def test_run_prospect_search_defaults_to_schemas_output_dir(tmp_path, monkeypatch):
    queries_csv = tmp_path / "queries.csv"
    queries_csv.write_text("google_search_query\nplumbers\n")

    output_dir = tmp_path / "data"
    output_dir.mkdir()
    monkeypatch.setattr(schemas, "OUTPUT_DIR", output_dir)
    monkeypatch.setattr(schemas, "BLACKLIST_PATH", _write_blacklist(tmp_path))

    def fake_http_get(url, headers, params, timeout):
        return FakeResponse({"web": {"results": [{"url": "https://example.com/1"}]}})

    dest_path = services.run_prospect_search(
        queries_csv,
        per_query=1,
        api_key="key",
        logger=logging.getLogger("test-run-default-dir"),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    assert dest_path.parent == output_dir / "results"


def test_run_prospect_search_filters_blacklisted_domains(tmp_path):
    queries_csv = tmp_path / "queries.csv"
    queries_csv.write_text("google_search_query\nshops\n")

    def fake_http_get(url, headers, params, timeout):
        results = [
            {"url": "https://allegro.pl/oferta/1"},
            {"url": "https://shop.allegro.pl/oferta/2"},
            {"url": "https://kept.example.com/page"},
        ]
        return FakeResponse({"web": {"results": results}})

    output_dir = tmp_path / "data"
    output_dir.mkdir()

    dest_path = services.run_prospect_search(
        queries_csv,
        per_query=3,
        api_key="key",
        logger=logging.getLogger("test-run-blacklist"),
        output_dir=output_dir,
        blacklist_path=_write_blacklist(tmp_path, ["allegro.pl"]),
        http_get=fake_http_get,
        sleep_fn=lambda seconds: None,
    )

    rows = list(load_workbook(dest_path).active.iter_rows(values_only=True))
    assert rows == [
        ("ID", "domain", "query"),
        (1, "kept.example.com", "shops"),
    ]


def test_run_prospect_search_with_no_queries_writes_header_only(tmp_path):
    queries_csv = tmp_path / "queries.csv"
    queries_csv.write_text("google_search_query\n")

    output_dir = tmp_path / "data"
    output_dir.mkdir()

    dest_path = services.run_prospect_search(
        queries_csv,
        per_query=5,
        api_key="key",
        logger=logging.getLogger("test-run-empty"),
        output_dir=output_dir,
        blacklist_path=_write_blacklist(tmp_path),
        http_get=lambda *args, **kwargs: pytest.fail("should not be called"),
        sleep_fn=lambda seconds: None,
    )

    rows = list(load_workbook(dest_path).active.iter_rows(values_only=True))
    assert rows == [("ID", "domain", "query")]


# --- Celery task ---


def test_run_prospect_search_task_calls_service(tmp_path, monkeypatch):
    captured = {}
    expected_dest = tmp_path / "results" / "20260101_000000.xlsx"

    def fake_run_prospect_search(**kwargs):
        captured.update(kwargs)
        return expected_dest

    monkeypatch.setattr(tasks.services, "run_prospect_search", fake_run_prospect_search)
    monkeypatch.setattr(
        tasks, "get_settings", lambda: SimpleNamespace(brave_api_key="task-key")
    )

    log_file = tmp_path / "run.log"
    queries_csv_path = tmp_path / "queries.csv"
    blacklist_path = tmp_path / "blacklist.txt"

    result = tasks.run_prospect_search_task(
        str(queries_csv_path), 10, str(log_file), str(tmp_path), str(blacklist_path)
    )

    assert result == str(expected_dest)
    assert captured["api_key"] == "task-key"
    assert captured["per_query"] == 10
    assert captured["queries_csv_path"] == queries_csv_path
    assert captured["output_dir"] == tmp_path
    assert captured["blacklist_path"] == blacklist_path
